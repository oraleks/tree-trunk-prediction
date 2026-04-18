"""
Trees per Capita Analysis

Joins Israel CBS population data with tree counts (both total and street
trees) to compute per-capita metrics. Also produces a correlation plot
between street trees per capita and mean Shade Index.

Usage:
    python tree_per_capita_analysis.py

Inputs:
    - CBS population file: d:/.../Data/Copy of אוכלוסייה לפי יישוב... .xlsx
    - urban_forest_data.xlsx (40 cities, n_trees)
    - street_trees_data.xlsx (18 cities, n_trees)
    - shade_index_data.xlsx (18 cities, mean_SI)

Outputs:
    - population_analysis.xlsx (3 sheets)
    - plots_shade_index/04_si_vs_street_trees_per_capita.png
"""

import os
import sys
import shutil
import tempfile
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.stats import pearsonr, spearmanr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings('ignore')

# =====================================================================
# Configuration
# =====================================================================

CBS_FILE = r"d:\OneDrive - Technion\Research\Shade Maps\Data\Copy of אוכלוסייה לפי יישוב, אזור סטטיסטי וגיל בודד - סוף 2023.xlsx"
URBAN_FOREST_XLSX = "urban_forest_data.xlsx"
STREET_TREES_XLSX = "street_trees_data.xlsx"
SHADE_INDEX_XLSX = "shade_index_data.xlsx"
OUTPUT_XLSX = "population_analysis.xlsx"
PLOT_DIR = "plots_shade_index"
PLOT_DIR_TPC = "plots_tree_per_capita"
CORR_PLOT = os.path.join(PLOT_DIR, "04_si_vs_street_trees_per_capita.png")
BAR_ALL_PLOT = os.path.join(PLOT_DIR_TPC, "01_trees_per_capita_all_cities.png")
BAR_STREET_PLOT = os.path.join(PLOT_DIR_TPC, "02_street_trees_per_capita.png")

# Mapping: city code -> (English name, Hebrew name in CBS file)
CITY_MAP = {
    'AFL': ('Afula', 'עפולה'),
    'AKO': ('Akko', 'עכו'),
    'ASD': ('Ashdod', 'אשדוד'),
    'ASK': ('Ashkelon', 'אשקלון'),
    'BBK': ('Bnei Brak', 'בני ברק'),
    'BSM': ('Beit Shemesh', 'בית שמש'),
    'BSV': ('Beersheva', 'באר שבע'),
    'BTR': ('Beitar Ilit', 'ביתר עילית'),
    'BTY': ('Bat Yam', 'בת ים'),
    'ELT': ('Eilat', 'אילת'),
    'GTM': ('Givatayim', 'גבעתיים'),
    'HAI': ('Haifa', 'חיפה'),
    'HDR': ('Hadera', 'חדרה'),
    'HDS': ('Hod HaSharon', 'הוד השרון'),
    'HOL': ('Holon', 'חולון'),
    'HRZ': ('Herzliya', 'הרצלייה'),
    'JER': ('Jerusalem', 'ירושלים'),
    'KAT': ('Kiryat Ata', 'קריית אתא'),
    'KFS': ('Kfar Saba', 'כפר סבא'),
    'KGT': ('Kiryat Gat', 'קריית גת'),
    'LOD': ('Lod', 'לוד'),
    'MDN': ('Modiin', 'מודיעין-מכבים-רעות'),
    'NHR': ('Nahariya', 'נהרייה'),
    'NSZ': ('Ness Ziona', 'נס ציונה'),
    'NTN': ('Netanya', 'נתניה'),
    'NTV': ('Netivot', 'נתיבות'),
    'NZR': ('Nazareth', 'נצרת'),
    'PHK': ('Pardes Hanna-Karkur', 'פרדס חנה-כרכור'),
    'PTV': ('Petah Tikva', 'פתח תקווה'),
    'RAN': ('Raanana', 'רעננה'),
    'RHT': ('Rahat', 'רהט'),
    'RHV': ('Rehovot', 'רחובות'),
    'RLZ': ('Rishon LeZion', 'ראשון לציון'),
    'RMG': ('Ramat Gan', 'רמת גן'),
    'RML': ('Ramla', 'רמלה'),
    'RSN': ('Rosh HaAyin', 'ראש העין'),
    'SDR': ('Sderot', 'שדרות'),
    'TLV': ('Tel Aviv', 'תל אביב -יפו'),
    'UMF': ('Umm al-Fahm', 'אום אל-פחם'),
    'YVN': ('Yavne', 'יבנה'),
}


# =====================================================================
# Data loading
# =====================================================================

def load_population():
    """Load CBS population file, return dict {hebrew_name: (cbs_code, total)}."""
    # Copy to temp location in case OneDrive file is locked
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    try:
        shutil.copy2(CBS_FILE, tmp.name)
        df = pd.read_excel(tmp.name, sheet_name='Sheet1', header=6)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    df = df[['סמל יישוב', 'שם יישוב', 'א"ס', 'סה"כ']].rename(
        columns={'סמל יישוב': 'cbs_code', 'שם יישוב': 'hebrew_name',
                 'א"ס': 'area', 'סה"כ': 'total'})

    # Keep only city-total rows (area == 'סה"כ')
    totals = df[df['area'] == 'סה"כ'].copy()
    # Strip whitespace in names and coerce code/total to numeric
    totals['hebrew_name'] = totals['hebrew_name'].astype(str).str.strip()
    totals['cbs_code'] = pd.to_numeric(totals['cbs_code'], errors='coerce')
    totals['total'] = pd.to_numeric(totals['total'], errors='coerce')
    totals = totals.dropna(subset=['cbs_code', 'total'])

    return {row['hebrew_name']: (int(row['cbs_code']), int(row['total']))
            for _, row in totals.iterrows()}


def load_tree_counts(xlsx_path):
    """Load n_trees per city from an analysis Excel file."""
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"{xlsx_path} not found -- run analyses first")
    df = pd.read_excel(xlsx_path, sheet_name='City Statistics')
    return dict(zip(df['city'], df['n_trees']))


def load_shade_index(xlsx_path):
    """Load mean_SI per city from shade_index_data.xlsx."""
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"{xlsx_path} not found -- run shade_index_analysis.py")
    df = pd.read_excel(xlsx_path, sheet_name='Shade Index')
    return dict(zip(df['city'], df['mean_SI']))


# =====================================================================
# Excel workbook creation
# =====================================================================

HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1A472A', end_color='1A472A', fill_type='solid')


def _format_headers(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ''
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def build_workbook(pop_map, all_trees, street_trees, output_path):
    """Build the multi-sheet population_analysis.xlsx."""
    wb = Workbook()

    # --- Sheet 1: Population All Cities ---
    ws1 = wb.active
    ws1.title = 'Population All Cities'
    headers = ['Code', 'English Name', 'Hebrew Name', 'CBS Code', 'Population']
    ws1.append(headers)

    missing = []
    sorted_cities = sorted(CITY_MAP.items(), key=lambda x: x[0])
    for code, (eng, heb) in sorted_cities:
        if heb not in pop_map:
            missing.append((code, eng, heb))
            ws1.append([code, eng, heb, None, None])
        else:
            cbs_code, pop = pop_map[heb]
            ws1.append([code, eng, heb, cbs_code, pop])

    _format_headers(ws1, len(headers))
    _autosize(ws1)

    if missing:
        print(f"\nWARNING: {len(missing)} cities could not be matched to CBS data:")
        for c, e, h in missing:
            print(f"  {c} ({e}) = {h!r}")

    # --- Sheet 2: Trees Per Capita - All Cities ---
    ws2 = wb.create_sheet('Trees Per Capita - All Cities')
    headers2 = ['Code', 'English Name', 'Hebrew Name', 'Population',
                'Total Trees', 'Trees per Capita']
    ws2.append(headers2)

    # Sort by n_trees descending
    rows_all = []
    for code, (eng, heb) in CITY_MAP.items():
        pop = pop_map.get(heb, (None, None))[1]
        n_trees = all_trees.get(code)
        rows_all.append((code, eng, heb, pop, n_trees))

    # Sort by trees per capita descending (skip missing)
    def tpc(row):
        _, _, _, pop, n = row
        if pop and n and pop > 0:
            return n / pop
        return -1
    rows_all.sort(key=tpc, reverse=True)

    for i, (code, eng, heb, pop, n_trees) in enumerate(rows_all, start=2):
        ws2.append([code, eng, heb, pop, n_trees, None])
        # Excel formula for trees per capita (column E / column D)
        ws2.cell(row=i, column=6).value = f"=IF(D{i}>0,E{i}/D{i},\"\")"
        ws2.cell(row=i, column=6).number_format = '0.0000'

    _format_headers(ws2, len(headers2))
    _autosize(ws2)

    # --- Sheet 3: Trees Per Capita - Streets ---
    ws3 = wb.create_sheet('Trees Per Capita - Streets')
    headers3 = ['Code', 'English Name', 'Hebrew Name', 'Population',
                'Street Trees', 'Street Trees per Capita']
    ws3.append(headers3)

    rows_street = []
    for code, (eng, heb) in CITY_MAP.items():
        if code not in street_trees:
            continue
        pop = pop_map.get(heb, (None, None))[1]
        n_street = street_trees.get(code)
        rows_street.append((code, eng, heb, pop, n_street))

    def stpc(row):
        _, _, _, pop, n = row
        if pop and n and pop > 0:
            return n / pop
        return -1
    rows_street.sort(key=stpc, reverse=True)

    for i, (code, eng, heb, pop, n_street) in enumerate(rows_street, start=2):
        ws3.append([code, eng, heb, pop, n_street, None])
        ws3.cell(row=i, column=6).value = f"=IF(D{i}>0,E{i}/D{i},\"\")"
        ws3.cell(row=i, column=6).number_format = '0.0000'

    _format_headers(ws3, len(headers3))
    _autosize(ws3)

    wb.save(output_path)
    print(f"Saved {output_path}")

    return len(rows_all), len(rows_street)


# =====================================================================
# Correlation plot
# =====================================================================

def plot_tpc_bar(pop_map, tree_counts, title_label, value_label, out_path):
    """Horizontal bar chart of trees-per-capita across cities, ranked."""
    records = []
    for code in tree_counts:
        eng, heb = CITY_MAP.get(code, (code, code))
        pop = pop_map.get(heb, (None, None))[1]
        if not pop or pop <= 0:
            continue
        n = tree_counts[code]
        records.append({
            'city': code,
            'city_name': eng,
            'population': pop,
            'n_trees': n,
            'trees_per_capita': n / pop,
        })

    if not records:
        print(f"  SKIP: no data for {title_label}")
        return None

    df = pd.DataFrame(records).sort_values('trees_per_capita')

    # Figure size scales with number of cities
    fig, ax = plt.subplots(figsize=(10, max(6, len(df) * 0.35)))

    # Color by tree-per-capita value (green gradient)
    cmap = plt.cm.YlGn
    norm = plt.Normalize(vmin=df['trees_per_capita'].min(),
                          vmax=df['trees_per_capita'].max())
    colors = [cmap(norm(v)) for v in df['trees_per_capita']]

    ax.barh(range(len(df)), df['trees_per_capita'], color=colors,
            edgecolor='black', linewidth=0.5)
    ax.set_yticks(range(len(df)))
    ax.set_yticklabels([f"{r['city']} ({r['city_name']})" for _, r in df.iterrows()],
                       fontsize=10)
    ax.set_xlabel(value_label, fontsize=12)
    ax.set_title(f'{title_label} (n={len(df)} cities)', fontsize=13)

    # Weighted mean line (by population)
    weighted_mean = np.average(df['trees_per_capita'], weights=df['population'])
    ax.axvline(weighted_mean, color='red', linestyle='--', linewidth=2,
               label=f'Population-weighted mean = {weighted_mean:.3f}')
    ax.legend(fontsize=11, loc='lower right')

    # Value labels
    for i, (_, row) in enumerate(df.iterrows()):
        ax.text(row['trees_per_capita'] + df['trees_per_capita'].max() * 0.005,
                i, f"{row['trees_per_capita']:.3f}",
                va='center', fontsize=9)

    ax.set_xlim(0, df['trees_per_capita'].max() * 1.15)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved {out_path}")

    return df


def plot_si_vs_street_tpc(pop_map, street_trees, shade_idx, out_path):
    """Correlation: street trees per capita vs mean SI (18 cities)."""
    records = []
    for code in street_trees:
        if code not in shade_idx:
            continue
        eng, heb = CITY_MAP.get(code, (code, code))
        pop = pop_map.get(heb, (None, None))[1]
        if not pop or pop <= 0:
            continue
        n_street = street_trees[code]
        records.append({
            'city': code,
            'city_name': eng,
            'street_trees': n_street,
            'population': pop,
            'trees_per_capita': n_street / pop,
            'mean_SI': shade_idx[code],
        })

    if not records:
        print("  SKIP: no overlapping cities for SI vs street-trees-per-capita")
        return None

    df = pd.DataFrame(records)

    r_pear, p_pear = pearsonr(df['trees_per_capita'], df['mean_SI'])
    r_spear, p_spear = spearmanr(df['trees_per_capita'], df['mean_SI'])

    fig, ax = plt.subplots(figsize=(9, 9))
    ax.scatter(df['trees_per_capita'], df['mean_SI'], s=80,
               c='forestgreen', edgecolors='black', linewidth=0.5,
               alpha=0.8, zorder=5)

    # Regression line
    z = np.polyfit(df['trees_per_capita'], df['mean_SI'], 1)
    xline = np.linspace(df['trees_per_capita'].min() * 0.95,
                         df['trees_per_capita'].max() * 1.05, 100)
    ax.plot(xline, np.polyval(z, xline), 'b-', linewidth=1.5, alpha=0.7,
            label=f'Fit: y={z[0]:.3f}x{z[1]:+.3f}', zorder=4)

    for _, row in df.iterrows():
        ax.annotate(row['city'], (row['trees_per_capita'], row['mean_SI']),
                    textcoords='offset points', xytext=(5, 5), fontsize=9)

    ax.set_xlabel('Street Trees per Capita (trees / resident)', fontsize=13)
    ax.set_ylabel('Mean Street Shade Index', fontsize=13)
    ax.set_title(f'Shade Index vs Street Trees per Capita (n={len(df)})\n'
                 f'Pearson r={r_pear:.3f} (p={p_pear:.4f}), '
                 f'Spearman rho={r_spear:.3f} (p={p_spear:.4f})',
                 fontsize=12)
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved {out_path}")

    return df, r_pear, r_spear


# =====================================================================
# Main
# =====================================================================

def main():
    os.makedirs(PLOT_DIR, exist_ok=True)
    os.makedirs(PLOT_DIR_TPC, exist_ok=True)

    print("=" * 60)
    print("Trees per Capita Analysis")
    print("=" * 60)

    print("\nLoading CBS population data...")
    pop_map = load_population()
    print(f"  Loaded population for {len(pop_map):,} localities")

    print("Loading tree counts...")
    all_trees = load_tree_counts(URBAN_FOREST_XLSX)
    street_trees = load_tree_counts(STREET_TREES_XLSX)
    print(f"  All trees: {len(all_trees)} cities")
    print(f"  Street trees: {len(street_trees)} cities")

    print("Loading shade index data...")
    shade_idx = load_shade_index(SHADE_INDEX_XLSX)
    print(f"  Shade index: {len(shade_idx)} cities")

    print("\nBuilding Excel workbook...")
    n_all, n_street = build_workbook(pop_map, all_trees, street_trees, OUTPUT_XLSX)
    print(f"  All cities sheet: {n_all} rows")
    print(f"  Street cities sheet: {n_street} rows")

    print("\nGenerating per-capita bar charts...")
    plot_tpc_bar(pop_map, all_trees,
                 'Trees per Capita by City (All Trees)',
                 'Trees per Capita',
                 BAR_ALL_PLOT)
    plot_tpc_bar(pop_map, street_trees,
                 'Street Trees per Capita by City',
                 'Street Trees per Capita',
                 BAR_STREET_PLOT)

    print("\nGenerating correlation plot...")
    result = plot_si_vs_street_tpc(pop_map, street_trees, shade_idx, CORR_PLOT)

    # Summary
    print("\n" + "=" * 60)
    print("Done.")
    print(f"  Workbook: {OUTPUT_XLSX}")
    print(f"  Plot: {CORR_PLOT}")
    if result:
        df, r_pear, r_spear = result
        print(f"\nStreet trees per capita vs mean SI (n={len(df)}):")
        print(f"  Pearson r = {r_pear:.3f}")
        print(f"  Spearman rho = {r_spear:.3f}")
        # Top and bottom
        df_sorted = df.sort_values('trees_per_capita', ascending=False)
        print(f"\nTop 5 street trees per capita:")
        for _, r in df_sorted.head(5).iterrows():
            print(f"  {r['city']} ({r['city_name']}): {r['trees_per_capita']:.3f} trees/capita, SI={r['mean_SI']:.3f}")
        print(f"\nBottom 5:")
        for _, r in df_sorted.tail(5).iterrows():
            print(f"  {r['city']} ({r['city_name']}): {r['trees_per_capita']:.3f} trees/capita, SI={r['mean_SI']:.3f}")


if __name__ == '__main__':
    main()
